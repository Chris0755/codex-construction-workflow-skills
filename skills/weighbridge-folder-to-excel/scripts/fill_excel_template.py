#!/usr/bin/env python3
"""Fill a copied Excel template from parsed rows without modifying the source."""
from __future__ import annotations

import argparse
import json
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl


DEFAULT_LAYOUT = {"date": 3, "material": 4, "category": 5, "unit": 6, "out_qty": 7, "in_qty": 8, "doc_no": 10, "car_no": 11, "site": 12, "supplier": 13, "team": 14, "note": 15, "tail_no": 16}


def load_json(path: Path | None, default):
    return json.loads(path.read_text(encoding="utf-8")) if path else default


def copy_format(source, target) -> None:
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def parse_date(value: str):
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except (TypeError, ValueError):
            continue
    return None


def load_plate_map(path: Path | None) -> dict[str, str]:
    if path is None or not path.exists():
        return {}
    mapping: dict[str, str] = {}
    workbook = openpyxl.load_workbook(path, data_only=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 2 and row[0] and row[1] and str(row[0]).strip() not in {"车牌", "车牌号", "车号"}:
                mapping[str(row[0]).strip().upper()] = str(row[1]).strip()
    return mapping


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", required=True, type=Path)
    parser.add_argument("--photo-folder", required=True, type=Path)
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--layout-config", type=Path)
    parser.add_argument("--plate-map", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--start-row", type=int, default=4)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    if not args.template.exists():
        raise FileNotFoundError(f"Template not found: {args.template}")
    rows = load_json(args.rows, [])
    layout_settings = load_json(args.layout_config, {})
    layout = {**DEFAULT_LAYOUT, **dict(layout_settings.get("columns") or layout_settings)}
    formula_column = layout_settings.get("formula_column")
    formula_template = layout_settings.get("formula_template")
    output = args.output or args.photo_folder / f"{args.template.stem}_{args.photo_folder.name}_filled{args.template.suffix}"
    source_mtime = args.template.stat().st_mtime
    workbook = openpyxl.load_workbook(args.template)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    plate_map = load_plate_map(args.plate_map)
    for index, row in enumerate(rows):
        target_row = args.start_row + index
        source_row = target_row if target_row <= sheet.max_row else args.start_row
        for column in range(1, sheet.max_column + 1):
            copy_format(sheet.cell(source_row, column), sheet.cell(target_row, column))
        values = {"date": parse_date(row.get("date")), "material": row.get("material") or "待核", "category": layout_settings.get("category_value", "aggregate"), "unit": layout_settings.get("unit_value", "t"), "out_qty": row.get("out_qty"), "in_qty": row.get("in_qty"), "doc_no": row.get("doc_no") or "待核", "car_no": row.get("car_no") or "待核", "site": row.get("site") or "待核", "supplier": row.get("supplier") or "", "team": plate_map.get(str(row.get("car_no") or "").strip().upper(), ""), "note": row.get("note") or "", "tail_no": row.get("tail_no") or ""}
        for name, value in values.items():
            sheet.cell(target_row, layout[name]).value = value
        if values["date"]:
            sheet.cell(target_row, layout["date"]).number_format = "yyyy/m/d"
        if formula_column and formula_template:
            sheet.cell(target_row, int(formula_column)).value = str(formula_template).format(row=target_row)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    if args.template.stat().st_mtime != source_mtime:
        raise RuntimeError("Template modified time changed")
    print(json.dumps({"output": str(output), "rows": len(rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
